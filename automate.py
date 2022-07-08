from openpyxl import load_workbook
from fuzzywuzzy import fuzz

#Utility functions

def changeActive(name, workbook):
	for i,v in enumerate(workbook.sheetnames):
		if v == name:
			workbook.active = i
			break


#Main
def main():
	wb_file = 'Customer General Code #2  6-28-2022.xlsx'

	wb = load_workbook(wb_file, data_only=True)

	changeActive('Sheet2', wb)
	sheet = wb.active

	i = 2

	#end at 2527

	#with open('test.txt', 'w') as f:
	while i <= 2527:
		count = 1

		for j in range(1, 11):
			one = sheet[f'E{i}'].value.replace('-', '').replace('_', '') if not sheet[f'E{i}'].value.split('-')[-1].isdigit() else sheet[f'E{i}'].value[:-4].replace('-', '').replace('_', '')
			two = sheet[f'E{i+j}'].value.replace('-', '').replace('_', '') if not sheet[f'E{i+j}'].value.split('-')[-1].isdigit() else sheet[f'E{i+j}'].value[:-4].replace('-', '').replace('_', '')
			onearr = ''.join([i for i in [char for char in one] if i.isdigit()])
			twoarr = ''.join([i for i in [char for char in two] if i.isdigit()])
			if fuzz.token_sort_ratio(one, two) >= 80 and fuzz.token_sort_ratio(onearr, twoarr) >= 85 or fuzz.token_sort_ratio(onearr, twoarr) >= 90:
				#f.write(f"{i}: {sheet[f'E{i}'].value}\n")
				sheet[f'D{i+j}'] = sheet[f'D{i}'].value
				print(i)
				count += 1
			else:
				sheet[f'D{i+j}'] = sheet[f'C{i+j}'].value
				break

		i += count


	wb.save(filename = wb_file)
	# tester1 = 'E2508'
	# tester2 = 'E2513'

	# one = sheet[tester1].value.replace('-', '').replace('_', '') if not sheet[tester1].value.split('-')[-1].isdigit() else sheet[tester1].value[:-4].replace('-', '').replace('_', '')
	# two = sheet[tester2].value.replace('-', '').replace('_', '') if not sheet[tester2].value.split('-')[-1].isdigit() else sheet[tester2].value[:-4].replace('-', '').replace('_', '')
	# print(one, two)

	# print(fuzz.token_sort_ratio(one, two))

	# onearr = ''.join([i for i in [char for char in one] if i.isdigit()])
	# twoarr = ''.join([i for i in [char for char in two] if i.isdigit()])
	# print(fuzz.token_sort_ratio(onearr, twoarr))
	# print(onearr, twoarr)


if __name__ == "__main__":
	main()